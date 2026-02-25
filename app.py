# =====================================================
# GYM COMMAND CENTRE
# =====================================================

import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth

# -----------------------------------------------------
# PAGE CONFIG
# -----------------------------------------------------
st.set_page_config(
    page_title="Gym Command Centre",
    layout="wide"
)

# =====================================================
# LOGIN SYSTEM
# =====================================================
with open("config.yaml") as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config["credentials"],
    config["cookie"]["name"],
    config["cookie"]["key"],
    config["cookie"]["expiry_days"]
)

authenticator.login(location="main")

authentication_status = st.session_state.get("authentication_status")
name = st.session_state.get("name")
username = st.session_state.get("username")

# =====================================================
# LOGIN CHECK
# =====================================================
if authentication_status is False:
    st.error("Username/password incorrect")

elif authentication_status is None:
    st.warning("Please login")

elif authentication_status:

    authenticator.logout("Logout","sidebar")
    st.sidebar.success(f"Welcome {name}")

    st.title("🏋️ Gym Command Centre")

    DATA_PATH="data"

    DECISION_FILE=os.path.join(
        DATA_PATH,"studio_decision_engine_v1.xlsx"
    )

    MONTH_FILE=os.path.join(
        DATA_PATH,"monthly_sessions.xlsx"
    )

    ATT_FILE=os.path.join(
        DATA_PATH,"attendance_log.xlsx"
    )

# =====================================================
# CLEAN NUMBER
# =====================================================
    def clean_number(val):

        if val is None:
            return 0

        if isinstance(val,(int,float)):
            if 0 < val < 1:
                return val*100
            return float(val)

        val=str(val).replace("₹","").replace(",","")

        try:
            num=float(val)
            if 0<num<1:
                num*=100
            return num
        except:
            return 0


# =====================================================
# KPI MAP
# =====================================================
    METRIC_MAP={
        "ACTIVE CLIENTS":"active",
        "MONTHLY REVENUE":"revenue",
        "REVENUE PER HOUR":"rev_hour",
        "AVERAGE UTILIZATION":"util",
        "MONTHLY PROFIT":"profit",
        "PROFIT MARGIN":"margin",
        "CAPACITY UTILIZATION VALUE":"cap_util",
        "REVENUE REALIZATION":"rev_real",
        "BREAK-EVEN CLIENT COUNT":"breakeven"
    }


    def extract_dashboard(sheet):

        results={}

        for row in sheet.iter_rows():
            for cell in row:

                if isinstance(cell.value,str):

                    label=cell.value.strip().upper()

                    if label in METRIC_MAP:

                        col=cell.column
                        start=cell.row+1

                        for r in range(start,start+6):

                            val=sheet.cell(r,col).value

                            if isinstance(val,(int,float)):
                                results[
                                    METRIC_MAP[label]
                                ]=clean_number(val)
                                break
        return results


# =====================================================
# KPI BOX UI (RESTORED)
# =====================================================
    def kpi_box(title,value):

        st.markdown(f"""
        <div style="
            background-color:#111827;
            padding:22px;
            border-radius:14px;
            border:1px solid #374151;
            text-align:center;">
            <h4 style="color:#9CA3AF;">
                {title}
            </h4>
            <h2 style="color:white;">
                {value}
            </h2>
        </div>
        """,unsafe_allow_html=True)


# =====================================================
# DATA LOADERS
# =====================================================
    def load_month():

        cols=["Client","Month","Total_Sessions"]

        if os.path.exists(MONTH_FILE):
            df=pd.read_excel(MONTH_FILE)
            if df.empty:
                df=pd.DataFrame(columns=cols)
        else:
            df=pd.DataFrame(columns=cols)

        return df


    def save_month(df):
        df.to_excel(MONTH_FILE,index=False)


    def load_att():

        cols=["Client","Month","Session_No","Status"]

        if os.path.exists(ATT_FILE):
            df=pd.read_excel(ATT_FILE)
            if df.empty:
                df=pd.DataFrame(columns=cols)
        else:
            df=pd.DataFrame(columns=cols)

        return df


    def save_att(df):
        df.to_excel(ATT_FILE,index=False)


# =====================================================
# NAVIGATION
# =====================================================
    pages=["Attendance","Month Setup"]

    if username=="aadhi":
        pages.insert(0,"Dashboard")

    page=st.sidebar.selectbox("Navigation",pages)

# =====================================================
# DASHBOARD ✅ FIXED
# =====================================================
    if page=="Dashboard":

        st.header("📊 Studio Dashboard")

        wb=load_workbook(
            DECISION_FILE,
            data_only=True
        )

        ws=wb["DASHBOARD"]
        m=extract_dashboard(ws)

        r1=st.columns(4)

        with r1[0]:
            kpi_box("Active Clients",int(m.get("active",0)))

        with r1[1]:
            kpi_box("Monthly Revenue",
                    f"₹{m.get('revenue',0):,.0f}")

        with r1[2]:
            kpi_box("Revenue / Hour",
                    f"₹{m.get('rev_hour',0):,.0f}")

        with r1[3]:
            kpi_box("Utilization",
                    f"{m.get('util',0):.0f}%")

        r2=st.columns(4)

        with r2[0]:
            kpi_box("Monthly Profit",
                    f"₹{m.get('profit',0):,.0f}")

        with r2[1]:
            kpi_box("Profit Margin",
                    f"{m.get('margin',0):.0f}%")

        with r2[2]:
            kpi_box("Capacity Utilization",
                    f"{m.get('cap_util',0):.0f}%")

        with r2[3]:
            kpi_box("Revenue Realization",
                    f"{m.get('rev_real',0):.0f}%")

        kpi_box("Break-even Clients",
                int(m.get("breakeven",0)))
